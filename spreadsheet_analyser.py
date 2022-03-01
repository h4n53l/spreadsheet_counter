import openpyxl
from easygui import fileopenbox, ynbox,textbox


def analyse_spreadsheet(filename):
    """
    :param filename: the path to the excel file
    :return: sheet_dimensions: A dictionary of row and column count per spreadsheet in the exsel file.
    """
    workbook = openpyxl.load_workbook(filename)

    # Dictionary to hold the number of rows and columns per spreadsheet.
    sheet_dimensions = {}

    # Get a list of spreadsheets in the excel file.
    all_sheets = workbook.sheetnames
    sheet_dimensions['List of Spreadsheets: '] = all_sheets

    for spreadsheet in all_sheets:
        # Get the total number of rows but do not include empty rows
        number_of_rows = len([row for row in workbook[spreadsheet] if not all([cell.value is None for cell in row])])

        # Get the total number of columns, empty or not.
        number_of_columns = workbook[spreadsheet].max_column

        # Create a key value pair for the total rows and columns in each spreadsheet.
        sheet_dimensions[f'Number of Rows in {spreadsheet}: '] = number_of_rows
        sheet_dimensions[f'Number of Columns in {spreadsheet}: '] = number_of_columns

    return sheet_dimensions

# This is not nicely written, but no one is going to worry about that.
if ynbox(msg='Ready to select an Excel file? ', title='Excel file parser'):
    excel_file = fileopenbox(msg='Select excel file', title='Excel Parser')
    file_info = analyse_spreadsheet(excel_file)
    total_rows = 0
    total_columns = 0

    texts = ''

    for key in file_info:
        if key.startswith('Number of Rows in'):
            total_rows += file_info[key]
        elif key.startswith('Number of Columns in'):
            total_columns += file_info[key]

    for line in file_info:
        texts += f"\n\t{line} \t{file_info[line]} \n"

    textbox(msg=f"This file has \t{len(file_info['List of Spreadsheets: '])} spreadsheets. \n" \
                f"Total Rows in all spreadsheets are: {total_rows} \n"
                f"Total Columns in all spreadsheets are: {total_columns}. "
            , title='Result',
            text=texts)

else:
    exit(0)
